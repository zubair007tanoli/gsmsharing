#!/usr/bin/env python3
"""
Layout-aware PDF to Excel conversion using pdfplumber + openpyxl.
Focus: keep rows/columns/tables readable and editable instead of dumping full-page text in one cell.
"""
import argparse
import os
import re
import sys
from collections import defaultdict

try:
    import pdfplumber
    import fitz
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    print(f"ERROR: Install required packages: pip install pdfplumber openpyxl. {exc}", file=sys.stderr)
    sys.exit(1)


THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def normalize_cell_text(value):
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def merge_positions(values, tolerance):
    if not values:
        return []
    values = sorted(values)
    merged = [values[0]]
    for v in values[1:]:
        if abs(v - merged[-1]) <= tolerance:
            merged[-1] = (merged[-1] + v) / 2.0
        else:
            merged.append(v)
    return merged


def choose_table_settings(mode):
    if mode == "fast":
        return {"vertical_strategy": "text", "horizontal_strategy": "text", "snap_tolerance": 4}
    if mode == "balanced":
        return {"vertical_strategy": "lines", "horizontal_strategy": "text", "snap_tolerance": 3}
    if mode == "exact":
        return {"vertical_strategy": "lines", "horizontal_strategy": "lines", "snap_tolerance": 2, "intersection_tolerance": 5}
    # accurate
    return {"vertical_strategy": "lines", "horizontal_strategy": "lines", "snap_tolerance": 2, "intersection_tolerance": 5}


def write_table(ws, start_row, table):
    row_idx = start_row
    max_col = 0
    for r_index, row in enumerate(table):
        values = [normalize_cell_text(c) for c in row]
        max_col = max(max_col, len(values))
        for c_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=11, bold=(r_index == 0))
            if r_index == 0 and value:
                cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
        row_idx += 1
    return row_idx, max_col


def infer_rows_from_words(page, mode, ai_enhance):
    words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False) or []
    if not words:
        return []

    y_bucket = 2 if mode == "accurate" else 3
    grouped = defaultdict(list)
    for w in words:
        top = float(w.get("top", 0))
        key = round(top / y_bucket) * y_bucket
        grouped[key].append(w)

    lines = []
    for key in sorted(grouped.keys()):
        line_words = sorted(grouped[key], key=lambda x: float(x.get("x0", 0)))
        lines.append(line_words)

    x_values = []
    for line in lines:
        for w in line:
            x_values.append(float(w.get("x0", 0)))

    tolerance = 18 if ai_enhance else 24
    anchors = merge_positions(x_values, tolerance=tolerance)
    if not anchors:
        anchors = [0.0]

    rows = []
    for line in lines:
        row = [""] * len(anchors)
        for w in line:
            x0 = float(w.get("x0", 0))
            text = normalize_cell_text(w.get("text", ""))
            if not text:
                continue
            col_idx = min(range(len(anchors)), key=lambda i: abs(anchors[i] - x0))
            if row[col_idx]:
                row[col_idx] = f"{row[col_idx]} {text}"
            else:
                row[col_idx] = text
        if any(cell.strip() for cell in row):
            rows.append(row)

    return rows


def autosize_columns(ws, max_widths):
    for col_idx, width in max_widths.items():
        if width <= 0:
            continue
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(10, width + 2), 60)


def build_exact_visual_workbook(input_pdf, output_xlsx):
    wb = Workbook()
    # remove default sheet; one sheet per page + hidden search index
    default_ws = wb.active
    wb.remove(default_ws)

    search_ws = wb.create_sheet("SearchIndex")
    search_ws.sheet_state = "hidden"
    search_ws.cell(row=1, column=1, value="Page")
    search_ws.cell(row=1, column=2, value="ExtractedText")

    temp_images = []

    with fitz.open(input_pdf) as doc:
        for page_num, page in enumerate(doc, start=1):
            ws = wb.create_sheet(f"Page_{page_num}")

            # render page at 2x for crisp fidelity
            matrix = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            img_path = f"{output_xlsx}.page_{page_num}.png"
            pix.save(img_path)
            temp_images.append(img_path)

            img = XLImage(img_path)
            ws.add_image(img, "A1")

            # tune sheet grid size to hold image area
            ws.column_dimensions["A"].width = max(10, int(pix.width / 7))
            ws.row_dimensions[1].height = max(15, int(pix.height * 0.75))

            # add searchable page text in hidden index sheet
            extracted_text = page.get_text("text") or ""
            search_ws.cell(row=page_num + 1, column=1, value=page_num)
            search_ws.cell(row=page_num + 1, column=2, value=extracted_text)

    out_dir = os.path.dirname(output_xlsx)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_xlsx)

    for p in temp_images:
        try:
            os.remove(p)
        except OSError:
            pass


def convert(input_pdf, output_xlsx, mode, ai_enhance):
    if not os.path.exists(input_pdf):
        raise FileNotFoundError(f"Input file not found: {input_pdf}")

    if mode == "exact":
        build_exact_visual_workbook(input_pdf, output_xlsx)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF_Extracted"

    current_row = 1
    max_widths = defaultdict(int)

    with pdfplumber.open(input_pdf) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            title = f"Page {page_num}"
            title_cell = ws.cell(row=current_row, column=1, value=title)
            title_cell.font = Font(name="Calibri", size=12, bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            current_row += 1

            page_data_written = False

            table_settings = choose_table_settings(mode)
            tables = page.extract_tables(table_settings=table_settings) or []
            for table in tables:
                cleaned = [[normalize_cell_text(c) for c in row] for row in table if row and any(normalize_cell_text(c) for c in row)]
                if not cleaned:
                    continue
                current_row, table_max_col = write_table(ws, current_row, cleaned)
                for row_vals in cleaned:
                    for c_idx, val in enumerate(row_vals, start=1):
                        max_widths[c_idx] = max(max_widths[c_idx], len(val))
                current_row += 1
                page_data_written = True

            if not page_data_written:
                inferred_rows = infer_rows_from_words(page, mode, ai_enhance)
                for row_vals in inferred_rows:
                    for c_idx, val in enumerate(row_vals, start=1):
                        cell = ws.cell(row=current_row, column=c_idx, value=val)
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.font = Font(name="Calibri", size=11)
                        max_widths[c_idx] = max(max_widths[c_idx], len(val))
                    current_row += 1
                page_data_written = bool(inferred_rows)

            if not page_data_written:
                fallback_text = normalize_cell_text(page.extract_text() or "")
                ws.cell(row=current_row, column=1, value=fallback_text or "[No extractable text found]")
                max_widths[1] = max(max_widths[1], len(fallback_text))
                current_row += 1

            current_row += 1

    autosize_columns(ws, max_widths)
    ws.freeze_panes = "A2"

    out_dir = os.path.dirname(output_xlsx)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_xlsx)


def parse_args():
    parser = argparse.ArgumentParser(description="Convert PDF to Excel with layout-aware extraction.")
    parser.add_argument("input_pdf")
    parser.add_argument("output_xlsx")
    parser.add_argument("--mode", choices=["exact", "accurate", "balanced", "fast"], default="exact")
    parser.add_argument("--ai-enhance", action="store_true", dest="ai_enhance")
    return parser.parse_args()


def main():
    args = parse_args()
    try:
        convert(args.input_pdf, args.output_xlsx, args.mode, args.ai_enhance)
        if os.path.exists(args.output_xlsx):
            print("SUCCESS")
            return 0
        print("ERROR: Output file was not created.", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
